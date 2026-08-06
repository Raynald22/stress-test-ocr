#!/usr/bin/env python3
"""
(dev/QA helper) Generate a corpus of "bad" Excel files for robustness.js.

The HPC upload path is: validate ext/size -> read .xlsx -> match sheets/columns
to a DB schema -> insert. Every one of those steps can meet a nasty file. This
script builds files that exercise each failure point, plus a bad_manifest.json
that robustness.js reads.

The rule robustness.js enforces is the same as the OCR suite's: a bad file must
be REJECTED CLEANLY (a proper 4xx, or at worst a handled 5xx) and must never
hang the request or crash the worker. Files that come back 200/"success" are
flagged as suspicious.

Usage:
    pip install openpyxl        # only needed for the structurally-valid cases
    python make_bad_xlsx.py

Output: ./samples/bad/*.xlsx  +  ./samples/bad/bad_manifest.json
Files that need openpyxl are skipped automatically if it isn't installed; the
byte-level bad files are always generated.
"""

import json
import os

OUT_DIR = os.path.join(os.path.dirname(__file__), "samples", "bad")
MAX_UPLOAD = 10 * 1024 * 1024  # keep in sync with MAX_UPLOAD_SIZE (default 10MB)

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

manifest = []


def add(fname, note, expect_reject=True):
    manifest.append({"file": f"bad/{fname}", "note": note, "expect_reject": expect_reject})


def write_bytes(fname, data):
    path = os.path.join(OUT_DIR, fname)
    with open(path, "wb") as f:
        f.write(data)
    return path


def main():
    os.makedirs(OUT_DIR, exist_ok=True)

    # 1) Not a zip at all — plain text renamed .xlsx. Parser should reject.
    write_bytes("not_really.xlsx", b"this is plainly not a spreadsheet\n" * 50)
    add("not_really.xlsx", "plain text renamed .xlsx (not a zip container)")

    # 2) Empty file (0 bytes).
    write_bytes("empty.xlsx", b"")
    add("empty.xlsx", "zero-byte file")

    # 3) Oversized (> MAX_UPLOAD) — should be rejected at the size gate (413)
    #    BEFORE any parsing. Content doesn't need to be valid xlsx.
    write_bytes("oversized.xlsx", b"\x00" * (MAX_UPLOAD + 512 * 1024))
    add("oversized.xlsx", "just over the 10MB upload limit -> expect 413")

    # The rest need a real xlsx container.
    try:
        from openpyxl import Workbook
    except ImportError:
        print("openpyxl not installed — skipping structurally-valid cases "
              "(truncated/blank/no_iddoc/many_rows). `pip install openpyxl` for full corpus.")
        _dump_manifest()
        return

    # Build one good workbook fully in memory to derive a truncated copy from
    # (BytesIO — no temp file to clean up, which OneDrive won't let us delete).
    import io
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["col_a", "col_b"])
    for i in range(20):
        ws.append([f"a{i}", f"b{i}"])
    buf = io.BytesIO()
    wb.save(buf)
    good_bytes = buf.getvalue()

    # 4) Truncated — valid header, cut off mid-file. Zip reader should error.
    write_bytes("truncated.xlsx", good_bytes[: len(good_bytes) // 2])
    add("truncated.xlsx", "valid xlsx truncated to half its bytes")

    # 5) Blank — structurally valid, single empty sheet, no data rows.
    wb = Workbook()
    wb.active.title = "Sheet1"
    wb.save(os.path.join(OUT_DIR, "blank.xlsx"))
    add("blank.xlsx", "valid xlsx but completely empty (no header, no rows)")

    # 6) Valid xlsx, real data, but NO id_doc / columns that map to no known DB
    #    schema. Exercises the schema-match failure, not the byte parser.
    wb = Workbook()
    ws = wb.active
    ws.title = "totally_unknown_sheet"
    ws.append(["nonsense_col_1", "nonsense_col_2", "nonsense_col_3"])
    for i in range(10):
        ws.append([i, f"x{i}", f"y{i}"])
    wb.save(os.path.join(OUT_DIR, "no_iddoc.xlsx"))
    add("no_iddoc.xlsx", "valid xlsx, but no id_doc / unmapped schema -> business reject")

    # 7) Many rows — structurally valid, heavy to parse (stress the parser).
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["col_a", "col_b", "col_c"])
    for i in range(50000):
        ws.append([f"a{i}", f"b{i}", i])
    wb.save(os.path.join(OUT_DIR, "many_rows.xlsx"))
    # This one may legitimately succeed OR fail depending on schema; the point is
    # it must not hang. Mark expect_reject False so a 200 isn't flagged, but the
    # no-hang / no-5xx checks still apply.
    add("many_rows.xlsx", "valid xlsx with 50k rows (parser load) — must not hang",
        expect_reject=False)

    _dump_manifest()


def _dump_manifest():
    # manifest lives one level up (samples/), file paths inside are "bad/<name>"
    path = os.path.join(os.path.dirname(OUT_DIR), "bad_manifest.json")
    with open(path, "w") as f:
        json.dump(manifest, f, indent=2)
    print(f"wrote {len(manifest)} bad files + manifest -> {path}")


if __name__ == "__main__":
    main()
